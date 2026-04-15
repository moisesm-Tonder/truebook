import io
from collections import defaultdict
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

MONTHS_ES_UPPER = [
    "ENERO",
    "FEBRERO",
    "MARZO",
    "ABRIL",
    "MAYO",
    "JUNIO",
    "JULIO",
    "AGOSTO",
    "SEPTIEMBRE",
    "OCTUBRE",
    "NOVIEMBRE",
    "DICIEMBRE",
]


def _num(value: Any) -> float:
    try:
        if value is None or value == "":
            return 0.0
        return float(value)
    except Exception:
        return 0.0


def _safe_str(value: Any, fallback: str = "") -> str:
    if value is None:
        return fallback
    return str(value)


def _parse_date(value: Any) -> Optional[date]:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if not value:
        return None
    s = str(value).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(s[:10], fmt).date()
        except Exception:
            continue
    return None


def _date_label(value: Any) -> str:
    d = _parse_date(value)
    if d:
        return d.isoformat()
    return _safe_str(value, "")


def _month_name_upper(period_month: int) -> str:
    idx = max(1, min(12, int(period_month))) - 1
    return MONTHS_ES_UPPER[idx]


def _styled_header_row(ws, row_idx: int, start_col: int, end_col: int):
    for col in range(start_col, end_col + 1):
        cell = ws.cell(row=row_idx, column=col)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F4E78")
        cell.alignment = Alignment(horizontal="center", vertical="center")


def _autowidth(ws, max_width: int = 52):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            value = "" if cell.value is None else str(cell.value)
            max_len = max(max_len, len(value))
        ws.column_dimensions[col_letter].width = min(max(10, max_len + 2), max_width)


def _save_workbook(wb: Workbook) -> bytes:
    buff = io.BytesIO()
    wb.save(buff)
    return buff.getvalue()


def _extract_fees_components(fees_result) -> Dict[str, Any]:
    merchant_summary = fees_result.merchant_summary or []
    daily_breakdown = fees_result.daily_breakdown or []
    withdrawals_summary = fees_result.withdrawals_summary or []
    refunds_summary = fees_result.refunds_summary or []
    other_fees_summary = fees_result.other_fees_summary or []
    return {
        "merchant_summary": merchant_summary,
        "daily_breakdown": daily_breakdown,
        "withdrawals_summary": withdrawals_summary,
        "refunds_summary": refunds_summary,
        "other_fees_summary": other_fees_summary,
    }


def build_fees_export(process, fees_result) -> Tuple[str, bytes]:
    period_text = f"{_month_name_upper(process.period_month)} {process.period_year}"
    components = _extract_fees_components(fees_result)
    merchant_summary = components["merchant_summary"]
    daily_breakdown = components["daily_breakdown"]
    withdrawals_summary = components["withdrawals_summary"]
    refunds_summary = components["refunds_summary"]
    other_fees_summary = components["other_fees_summary"]

    wb = Workbook()
    ws_detail = wb.active
    ws_detail.title = "Detalle por Merchant"
    ws_summary = wb.create_sheet("Resumen por Merchant")
    ws_razon = wb.create_sheet("Resumen por Razon Social")
    ws_daily = wb.create_sheet("Tonder Fees desglose diario")

    # ---- Detalle por Merchant ----
    ws_detail["A2"] = f"REPORTE DE FEES — {period_text}   |   DETALLE POR MERCHANT"
    ws_detail["A3"] = (
        f"Montos en MXN  |  UTC-6  |  01-{_month_name_upper(process.period_month)[:3].title()}-{process.period_year} "
        f"00:00 → Fin de mes 23:59  |  Neto = Monto Procesado - Fee c/IVA"
    )
    detail_headers = [
        "",
        "Merchant",
        "Concepto",
        "Adquirente",
        "# Eventos",
        "Monto Procesado",
        "Fee %",
        "Fee Fijo",
        "Total Fee s/IVA",
        "IVA (16%)",
        "Total c/IVA",
        "Neto a Liquidar",
    ]
    ws_detail.append([])
    ws_detail.append(detail_headers)
    _styled_header_row(ws_detail, ws_detail.max_row, 1, len(detail_headers))

    grouped_tx = defaultdict(lambda: {"events": 0, "amount": 0.0, "fee": 0.0})
    merchant_name_map: Dict[str, str] = {}
    for row in daily_breakdown:
        merchant_id = _safe_str(row.get("merchant_id"), "unknown")
        merchant_name = _safe_str(row.get("merchant_name"), merchant_id)
        acquirer = _safe_str(row.get("acquirer"), "Operativa")
        key = (merchant_id, merchant_name, f"{acquirer} - Operativa", acquirer)
        grouped_tx[key]["events"] += 1
        grouped_tx[key]["amount"] += _num(row.get("amount"))
        grouped_tx[key]["fee"] += _num(row.get("fee_amount"))
        merchant_name_map[merchant_id] = merchant_name

    w_map = { _safe_str(w.get("merchant_id"), "unknown"): w for w in withdrawals_summary }
    r_map = { _safe_str(r.get("merchant_id"), "unknown"): r for r in refunds_summary }

    all_merchants = set(merchant_name_map.keys()) | set(w_map.keys()) | set(r_map.keys())
    all_merchants |= { _safe_str(m.get("merchant_id"), "unknown") for m in merchant_summary }

    for mid in sorted(all_merchants, key=lambda x: merchant_name_map.get(x, x)):
        merchant_name = merchant_name_map.get(mid, mid)
        ws_detail.append([None, merchant_name])

        for (g_mid, _m_name, concept, acquirer), val in sorted(grouped_tx.items(), key=lambda x: (x[0][1], x[0][2])):
            if g_mid != mid:
                continue
            amount = round(val["amount"], 6)
            total_fee = round(val["fee"], 6)
            fee_pct = round((total_fee / amount) if amount else 0.0, 6)
            iva = round(total_fee * 0.16, 6)
            total_c_iva = round(total_fee + iva, 6)
            neto = round(amount - total_c_iva, 6)
            ws_detail.append([
                None,
                None,
                concept,
                acquirer,
                val["events"],
                amount,
                fee_pct,
                0.0,
                total_fee,
                iva,
                total_c_iva,
                neto,
            ])

        if mid in w_map:
            w = w_map[mid]
            total_fee = round(_num(w.get("total_fee")), 6)
            amount = round(_num(w.get("total_amount")), 6)
            iva = round(total_fee * 0.16, 6)
            ws_detail.append([
                None,
                None,
                "Withdrawals",
                "N/A",
                int(_num(w.get("count"))),
                amount if amount else "—",
                round((total_fee / amount) if amount else 0.0, 6),
                0.0,
                total_fee,
                iva,
                round(total_fee + iva, 6),
                "—",
            ])

        if mid in r_map:
            r = r_map[mid]
            total_fee = round(_num(r.get("total_fee")), 6)
            amount = round(_num(r.get("total_amount")), 6)
            iva = round(total_fee * 0.16, 6)
            ws_detail.append([
                None,
                None,
                "Autorefunds/Refunds",
                "N/A",
                int(_num(r.get("count"))),
                amount if amount else "—",
                round((total_fee / amount) if amount else 0.0, 6),
                0.0,
                total_fee,
                iva,
                round(total_fee + iva, 6),
                "—",
            ])

    # ---- Resumen por Merchant ----
    ws_summary["A2"] = f"RESUMEN DE FEES POR MERCHANT — {period_text}"
    ws_summary["A3"] = "Montos en MXN  |  Neto = Monto Procesado - Fee c/IVA"
    ws_summary.append([])
    summary_headers = [
        "",
        "Merchant",
        "Monto Procesado",
        "Fees Transacc.",
        "Other Fees",
        "Settlement",
        "Withdrawals",
        "Autorefunds",
        "Routing Fee",
        "Total s/IVA",
        "IVA (16%)",
        "Total c/IVA",
        "Neto a Liquidar",
    ]
    ws_summary.append(summary_headers)
    _styled_header_row(ws_summary, ws_summary.max_row, 1, len(summary_headers))

    other_by_mid = defaultdict(float)
    settlement_by_mid = defaultdict(float)
    routing_by_mid = defaultdict(float)
    for item in other_fees_summary if isinstance(other_fees_summary, list) else []:
        mid = _safe_str(item.get("merchant_id"), "unknown")
        amount = _num(item.get("total_fee", item.get("amount", 0)))
        concept = _safe_str(item.get("concept", item.get("type", "other"))).lower()
        if "settlement" in concept:
            settlement_by_mid[mid] += amount
        elif "routing" in concept:
            routing_by_mid[mid] += amount
        else:
            other_by_mid[mid] += amount

    merchant_ids = [ _safe_str(m.get("merchant_id"), "unknown") for m in merchant_summary ]
    merchant_rows = { _safe_str(m.get("merchant_id"), "unknown"): m for m in merchant_summary }
    for mid in sorted(set(merchant_ids) | set(w_map.keys()) | set(r_map.keys())):
        m = merchant_rows.get(mid, {})
        name = _safe_str(m.get("merchant_name"), mid)
        gross = round(_num(m.get("gross_amount")), 6)
        tx_fee = round(_num(m.get("total_fee")), 6)
        withdrawals = round(_num((w_map.get(mid) or {}).get("total_fee")), 6)
        autorefunds = round(_num((r_map.get(mid) or {}).get("total_fee")), 6)
        other_fees = round(other_by_mid[mid], 6)
        settlement = round(settlement_by_mid[mid], 6)
        routing = round(routing_by_mid[mid], 6)
        total_s_iva = round(tx_fee + other_fees + settlement + withdrawals + autorefunds + routing, 6)
        iva = round(total_s_iva * 0.16, 6)
        total_c_iva = round(total_s_iva + iva, 6)
        neto = round(gross - total_c_iva, 6) if gross else "—"
        ws_summary.append([
            None,
            name,
            gross,
            tx_fee,
            other_fees,
            settlement,
            withdrawals,
            autorefunds,
            routing if routing else "—",
            total_s_iva,
            iva,
            total_c_iva,
            neto,
        ])

    # ---- Resumen por Razon Social ----
    ws_razon["A2"] = f"RESUMEN DE FEES POR RAZON SOCIAL — {period_text}"
    ws_razon["A3"] = "Montos en MXN  |  Consolidado por razón social (placeholder por merchant)"
    ws_razon.append([])
    razon_headers = [
        "",
        "Razon Social",
        "Merchants",
        "Monto Procesado",
        "Fees Transacc.",
        "Other Fees",
        "Settlement",
        "Withdrawals",
        "Routing Fee",
        "Total s/IVA",
        "IVA (16%)",
        "Total c/IVA",
        "Neto a Liquidar",
    ]
    ws_razon.append(razon_headers)
    _styled_header_row(ws_razon, ws_razon.max_row, 1, len(razon_headers))

    for row in ws_summary.iter_rows(min_row=6, values_only=True):
        if not row or not row[1]:
            continue
        ws_razon.append([
            None,
            row[1],  # Razon Social (sin catálogo aún)
            row[1],  # Merchants
            row[2],
            row[3],
            row[4],
            row[5],
            row[6],
            row[8],
            row[9],
            row[10],
            row[11],
            row[12],
        ])

    # ---- Desglose diario ----
    ws_daily["A1"] = (
        f"TONDER — FEES {period_text} | DESGLOSE DIARIO | Montos en MXN | UTC-6 | "
        f"01-{_month_name_upper(process.period_month)[:3].title()}-{process.period_year} → Fin de mes"
    )
    daily_headers = [
        "Fecha",
        "Merchant",
        "Concepto / Operativa",
        "# Eventos",
        "Monto Procesado",
        "Fee s/IVA",
        "IVA (16%)",
        "Total c/IVA",
    ]
    ws_daily.append(daily_headers)
    _styled_header_row(ws_daily, 2, 1, len(daily_headers))

    daily_grouped = defaultdict(lambda: {"events": 0, "amount": 0.0, "fee": 0.0})
    for row in daily_breakdown:
        d = _date_label(row.get("date"))
        merchant = _safe_str(row.get("merchant_name"), _safe_str(row.get("merchant_id"), "unknown"))
        acquirer = _safe_str(row.get("acquirer"), "Operativa")
        concept = f"{acquirer}-Operativa"
        key = (d, merchant, concept)
        daily_grouped[key]["events"] += 1
        daily_grouped[key]["amount"] += _num(row.get("amount"))
        daily_grouped[key]["fee"] += _num(row.get("fee_amount"))

    last_date = None
    for (d, merchant, concept), val in sorted(daily_grouped.items(), key=lambda x: (x[0][0], x[0][1], x[0][2])):
        if d != last_date:
            pretty = d
            parsed = _parse_date(d)
            if parsed:
                pretty = parsed.strftime("  %d/%m/%Y")
            ws_daily.append([pretty])
            last_date = d
        fee = round(val["fee"], 6)
        iva = round(fee * 0.16, 6)
        ws_daily.append([
            None,
            merchant,
            concept,
            val["events"],
            round(val["amount"], 6),
            fee,
            iva,
            round(fee + iva, 6),
        ])

    for ws in [ws_detail, ws_summary, ws_razon, ws_daily]:
        _autowidth(ws)

    filename = f"FEES_{_month_name_upper(process.period_month)}_{process.period_year}_FINAL.xlsx"
    return filename, _save_workbook(wb)


def build_kushki_export(process, kushki_result) -> Tuple[str, bytes]:
    period_text = f"{_month_name_upper(process.period_month)} {process.period_year}"
    daily_summary = sorted((kushki_result.daily_summary or []), key=lambda r: _date_label(r.get("date")))
    merchant_detail = sorted(
        (kushki_result.merchant_detail or []),
        key=lambda r: (_date_label(r.get("date")), _safe_str(r.get("merchant_name"), "").lower()),
    )

    wb = Workbook()
    ws_detail = wb.active
    ws_detail.title = "Detalle por Merchant"
    ws_daily = wb.create_sheet("Resumen Diario")
    ws_pivot = wb.create_sheet("Pivot por Merchant")

    teal = "1A7A6E"
    teal_light = "E8F4F2"
    teal_med = "A8D5CF"
    teal_dark = "145F55"

    def style_header(ws, row_idx: int, start_col: int, end_col: int):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=teal)
            cell.alignment = Alignment(horizontal="center", vertical="center")

    def style_total(ws, row_idx: int, start_col: int, end_col: int):
        for col in range(start_col, end_col + 1):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = PatternFill("solid", fgColor=teal_dark)

    def style_block_header(ws, row_idx: int):
        for col in range(1, 20):
            cell = ws.cell(row=row_idx, column=col)
            cell.font = Font(bold=True, color="0F172A")
            cell.fill = PatternFill("solid", fgColor=teal_med)

    def apply_number_formats(ws, start_row: int, end_row: int, count_cols: List[int], amount_cols: List[int], pct_cols: List[int]):
        for r in range(start_row, end_row + 1):
            for c in count_cols:
                ws.cell(row=r, column=c).number_format = "#,##0"
            for c in amount_cols:
                ws.cell(row=r, column=c).number_format = "#,##0.00"
            for c in pct_cols:
                ws.cell(row=r, column=c).number_format = "0.00%"

    daily_by_date = {str(row.get("date", "")): row for row in daily_summary}
    merchant_by_date: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
    for row in merchant_detail:
        merchant_by_date[str(row.get("date", ""))].append(row)

    # ---- Detalle por Merchant ----
    ws_detail["A1"] = f"KUSHKI — DETALLE POR MERCHANT · {period_text}"
    ws_detail["A2"] = "Montos en MXN | Estructura contable por liquidación (settlement)"
    detail_headers = [
        "Fecha Liq.",
        "Merchant",
        "# Txns",
        "Monto Bruto (Kushki)",
        "Bruto Ajustes",
        "Com. Kushki",
        "IVA Kushki",
        "Com. Kushki + IVA",
        "RR Retenido",
        "Devolución (REFUND)",
        "Contracargo (CHARGEBACK)",
        "Cancelación (VOID)",
        "Manual (MANUAL)",
        "RR Liberado",
        "Depósito Neto (Monto Abonar)",
        "Com. Tonder s/IVA",
        "IVA (16%)",
        "Com. Tonder c/IVA",
        "Verificación",
    ]
    ws_detail.append(detail_headers)
    style_header(ws_detail, 3, 1, len(detail_headers))
    ws_detail.freeze_panes = "A4"

    first_detail_data_row = None
    last_detail_data_row = None

    for idx, liq_date in enumerate(sorted(set(list(daily_by_date.keys()) + list(merchant_by_date.keys())))):
        if not liq_date:
            continue
        daily = daily_by_date.get(liq_date, {})
        dep_total = round(_num(daily.get("net_deposit")), 6)

        ws_detail.append([f"Liquidación {liq_date} | Depósito Banregio esperado: ${dep_total:,.2f}"])
        block_row = ws_detail.max_row
        ws_detail.merge_cells(start_row=block_row, start_column=1, end_row=block_row, end_column=19)
        style_block_header(ws_detail, block_row)

        merchant_rows = merchant_by_date.get(liq_date, [])
        merchant_rr_total = 0.0
        for row in merchant_rows:
            rr_released = round(_num(row.get("rr_released")), 6)
            merchant_rr_total += rr_released
            ws_detail.append([
                liq_date,
                _safe_str(row.get("merchant_name"), "unknown"),
                int(_num(row.get("tx_count"))),
                round(_num(row.get("gross_amount")), 6),
                round(_num(row.get("gross_adjustments")), 6),
                round(_num(row.get("kushki_commission")), 6),
                round(_num(row.get("iva_kushki_commission")), 6),
                round(_num(row.get("commission")), 6),
                round(_num(row.get("rr_retained")), 6),
                round(_num(row.get("refund")), 6),
                round(_num(row.get("chargeback")), 6),
                round(_num(row.get("void")), 6),
                round(_num(row.get("manual")), 6),
                rr_released,
                round(_num(row.get("net_deposit")), 6),
                round(_num(row.get("tonder_commission")), 6),
                round(_num(row.get("tonder_iva")), 6),
                round(_num(row.get("tonder_total")), 6),
                round(_num(row.get("net_deposit")), 6),
            ])
            current_row = ws_detail.max_row
            if first_detail_data_row is None:
                first_detail_data_row = current_row
            last_detail_data_row = current_row
            if (idx + current_row) % 2 == 0:
                for col in range(1, 20):
                    ws_detail.cell(row=current_row, column=col).fill = PatternFill("solid", fgColor=teal_light)

        rr_released_daily = round(_num(daily.get("rr_released")), 6)
        rr_released_extra = round(rr_released_daily - merchant_rr_total, 6)
        if abs(rr_released_extra) > 0.01:
            ws_detail.append([
                liq_date,
                "RR Liberado (Ajuste de bloque)",
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                None,
                rr_released_extra,
                rr_released_extra,
                None,
                None,
                None,
                rr_released_extra,
            ])
            current_row = ws_detail.max_row
            if first_detail_data_row is None:
                first_detail_data_row = current_row
            last_detail_data_row = current_row

        ws_detail.append([])

    ws_detail.append(["TOTAL"])
    detail_total_row = ws_detail.max_row
    if first_detail_data_row and last_detail_data_row:
        for col in range(3, 20):
            col_letter = get_column_letter(col)
            ws_detail.cell(
                row=detail_total_row,
                column=col,
                value=f"=SUM({col_letter}{first_detail_data_row}:{col_letter}{last_detail_data_row})",
            )
    style_total(ws_detail, detail_total_row, 1, 19)

    # ---- Resumen Diario ----
    ws_daily["A1"] = f"KUSHKI — RESUMEN DIARIO DE LIQUIDACIONES · {period_text}"
    daily_headers = [
        "Fecha Liq.",
        "# Txns",
        "Monto Bruto (Kushki)",
        "Bruto Ajustes",
        "Com. Kushki + IVA",
        "RR Retenido",
        "Devolución (REFUND)",
        "Contracargo (CHARGEBACK)",
        "Cancelación (VOID)",
        "Manual (MANUAL)",
        "RR Liberado",
        "Depósito Neto",
        "Com. Tonder s/IVA",
        "IVA (16%)",
        "Com. Tonder c/IVA",
    ]
    ws_daily.append(daily_headers)
    style_header(ws_daily, 2, 1, len(daily_headers))
    ws_daily.freeze_panes = "A3"

    daily_start_row = ws_daily.max_row + 1
    for row in daily_summary:
        ws_daily.append([
            _date_label(row.get("date")),
            int(_num(row.get("tx_count"))),
            round(_num(row.get("gross_amount")), 6),
            round(_num(row.get("gross_adjustments")), 6),
            round(_num(row.get("commission")), 6),
            round(_num(row.get("rr_retained")), 6),
            round(_num(row.get("refund")), 6),
            round(_num(row.get("chargeback")), 6),
            round(_num(row.get("void")), 6),
            round(_num(row.get("manual")), 6),
            round(_num(row.get("rr_released")), 6),
            round(_num(row.get("net_deposit")), 6),
            round(_num(row.get("tonder_commission")), 6),
            round(_num(row.get("tonder_iva")), 6),
            round(_num(row.get("tonder_total")), 6),
        ])

    ws_daily.append(["TOTAL"])
    daily_total_row = ws_daily.max_row
    if ws_daily.max_row > daily_start_row:
        for col in range(2, 16):
            col_letter = get_column_letter(col)
            ws_daily.cell(
                row=daily_total_row,
                column=col,
                value=f"=SUM({col_letter}{daily_start_row}:{col_letter}{daily_total_row - 1})",
            )
    style_total(ws_daily, daily_total_row, 1, 15)

    # ---- Pivot por Merchant ----
    ws_pivot["A1"] = f"KUSHKI — ACUMULADO POR MERCHANT · {period_text}"
    pivot_headers = [
        "Merchant",
        "# Txns",
        "Monto Bruto (Kushki)",
        "Com. Kushki + IVA",
        "RR Retenido",
        "Tasa Efectiva",
        "Ajuste Total",
        "Depósito Neto",
    ]
    ws_pivot.append(pivot_headers)
    style_header(ws_pivot, 2, 1, len(pivot_headers))
    ws_pivot.freeze_panes = "A3"

    pivot_acc = defaultdict(
        lambda: {
            "tx_count": 0.0,
            "gross_amount": 0.0,
            "commission": 0.0,
            "rr_retained": 0.0,
            "adjustment_total": 0.0,
            "net_deposit": 0.0,
        }
    )
    for row in merchant_detail:
        merchant = _safe_str(row.get("merchant_name"), "unknown")
        pivot_acc[merchant]["tx_count"] += _num(row.get("tx_count"))
        pivot_acc[merchant]["gross_amount"] += _num(row.get("gross_amount"))
        pivot_acc[merchant]["commission"] += _num(row.get("commission"))
        pivot_acc[merchant]["rr_retained"] += _num(row.get("rr_retained"))
        pivot_acc[merchant]["adjustment_total"] += (
            _num(row.get("refund"))
            + _num(row.get("chargeback"))
            + _num(row.get("void"))
            + _num(row.get("manual"))
            + _num(row.get("rr_released"))
        )
        pivot_acc[merchant]["net_deposit"] += _num(row.get("net_deposit"))

    pivot_start_row = ws_pivot.max_row + 1
    for merchant, row in sorted(pivot_acc.items(), key=lambda x: x[0].lower()):
        gross = round(_num(row.get("gross_amount")), 6)
        commission = round(_num(row.get("commission")), 6)
        tasa = (commission / gross) if gross else 0.0
        ws_pivot.append([
            merchant,
            int(_num(row.get("tx_count"))),
            gross,
            commission,
            round(_num(row.get("rr_retained")), 6),
            tasa,
            round(_num(row.get("adjustment_total")), 6),
            round(_num(row.get("net_deposit")), 6),
        ])

    ws_pivot.append(["TOTAL"])
    pivot_total_row = ws_pivot.max_row
    if ws_pivot.max_row > pivot_start_row:
        for col in [2, 3, 4, 5, 7, 8]:
            col_letter = get_column_letter(col)
            ws_pivot.cell(
                row=pivot_total_row,
                column=col,
                value=f"=SUM({col_letter}{pivot_start_row}:{col_letter}{pivot_total_row - 1})",
            )
    style_total(ws_pivot, pivot_total_row, 1, 8)

    # Column widths (spec-recommended)
    detail_widths = [12, 26, 8, 16, 14, 12, 12, 14, 13, 13, 15, 13, 11, 13, 16, 15, 13, 15, 13]
    daily_widths = [12, 10, 16, 14, 16, 13, 14, 16, 13, 11, 13, 15, 15, 13, 15]
    pivot_widths = [28, 10, 18, 16, 14, 14, 14, 16]
    for idx, width in enumerate(detail_widths, start=1):
        ws_detail.column_dimensions[get_column_letter(idx)].width = width
    for idx, width in enumerate(daily_widths, start=1):
        ws_daily.column_dimensions[get_column_letter(idx)].width = width
    for idx, width in enumerate(pivot_widths, start=1):
        ws_pivot.column_dimensions[get_column_letter(idx)].width = width

    # Number formatting
    if first_detail_data_row and detail_total_row >= first_detail_data_row:
        apply_number_formats(
            ws_detail,
            start_row=first_detail_data_row,
            end_row=detail_total_row,
            count_cols=[3],
            amount_cols=list(range(4, 20)),
            pct_cols=[],
        )
    if daily_start_row and daily_total_row >= daily_start_row:
        apply_number_formats(
            ws_daily,
            start_row=daily_start_row,
            end_row=daily_total_row,
            count_cols=[2],
            amount_cols=list(range(3, 16)),
            pct_cols=[],
        )
    if pivot_start_row and pivot_total_row >= pivot_start_row:
        apply_number_formats(
            ws_pivot,
            start_row=pivot_start_row,
            end_row=pivot_total_row,
            count_cols=[2],
            amount_cols=[3, 4, 5, 7, 8],
            pct_cols=[6],
        )

    filename = f"KUSHKI_{_month_name_upper(process.period_month)}_{process.period_year}_v4.xlsx"
    return filename, _save_workbook(wb)


def _amount_key(value: Any) -> int:
    return int(round(_num(value) * 100, 0))


def _infer_category(description: str, debit: float, credit: float, is_kushki_credit: bool = False) -> str:
    d = (description or "").lower()
    if is_kushki_credit or "kushki" in d:
        return "Kushki - Liquidacion"
    if "settlement" in d or ("stp" in d and debit > 0):
        return "Settlement merchant"
    if "spei" in d and credit > 0:
        return "Abonos SPEI"
    if "spei" in d and debit > 0:
        return "Cargos SPEI"
    if credit > 0:
        return "Abonos"
    if debit > 0:
        return "Cargos"
    return "Otros"


def build_banregio_export(process, banregio_result, kushki_result, conciliation_results: List[Any]) -> Tuple[str, bytes]:
    period_text = f"{_month_name_upper(process.period_month)} {process.period_year}"
    movements = banregio_result.movements or []
    summary = banregio_result.summary or {}
    daily_summary = (kushki_result.daily_summary if kushki_result else []) or []

    kvb = None
    for c in conciliation_results or []:
        if getattr(c, "conciliation_type", "") == "kushki_vs_banregio":
            kvb = c
            break

    matched = (kvb.matched if kvb else []) or []
    kvb_differences = (kvb.differences if kvb else []) or []
    unmatched_banregio = (kvb.unmatched_banregio if kvb else []) or []
    total_difference = _num(getattr(kvb, "total_difference", 0)) if kvb else 0.0

    # Lookup to identify Banregio credits related to Kushki deposits.
    kushki_credit_by_date = defaultdict(set)
    kushki_credit_any_date = set()
    for row in daily_summary:
        amount = _num(row.get("net_deposit"))
        if amount <= 0:
            continue
        date_key = _date_label(row.get("date"))
        amount_key = _amount_key(amount)
        if date_key:
            kushki_credit_by_date[date_key].add(amount_key)
        kushki_credit_any_date.add(amount_key)
    for row in (matched + kvb_differences):
        amount = _num(row.get("banregio_amount"))
        if amount <= 0:
            continue
        date_key = _date_label(row.get("banregio_date") or row.get("date"))
        amount_key = _amount_key(amount)
        if date_key:
            kushki_credit_by_date[date_key].add(amount_key)
        kushki_credit_any_date.add(amount_key)

    wb = Workbook()
    ws_moves = wb.active
    ws_moves.title = "MOVIMIENTOS"
    ws_summary = wb.create_sheet("RESUMEN")
    ws_cross = wb.create_sheet("CRUCE KUSHKI")

    # ---- MOVIMIENTOS ----
    ws_moves["A1"] = f"ESTADO DE CUENTA BANREGIO – {period_text}   |   TRES COMAS S.A.P.I. DE C.V.   |   Cta. 001-9"
    ws_moves["A2"] = (
        f"Total Abonos: ${_num(summary.get('total_credits')):,.2f}  |  "
        f"Total Cargos: ${_num(summary.get('total_debits')):,.2f}  |  "
        f"Neto: ${_num(summary.get('net')):,.2f}"
    )
    move_headers = ["DÍA", "FECHA", "TIPO", "CONCEPTO / CONTRAPARTE", "CATEGORÍA", "CARGOS (MXN)", "ABONOS (MXN)", "SALDO (MXN)"]
    ws_moves.append(move_headers)
    _styled_header_row(ws_moves, 3, 1, len(move_headers))

    category_acc = defaultdict(lambda: {"debits": 0.0, "credits": 0.0, "count": 0})
    running_balance = 0.0
    for mv in movements:
        d = _parse_date(mv.get("date"))
        debit = round(_num(mv.get("debit")), 6)
        credit = round(_num(mv.get("credit")), 6)
        desc = _safe_str(mv.get("description"), "")
        typ = _safe_str(mv.get("type"), "INT" if ("spei" in desc.lower() or "traspaso" in desc.lower()) else "")
        date_key = _date_label(mv.get("date"))
        credit_key = _amount_key(credit)
        desc_lower = desc.lower()
        is_kushki_credit = (
            credit > 0
            and (
                "kushki" in desc_lower
                or (date_key and credit_key in kushki_credit_by_date.get(date_key, set()))
                or credit_key in kushki_credit_any_date
            )
        )
        category = _safe_str(mv.get("category"), "").strip()
        if is_kushki_credit:
            category = "Kushki - Liquidacion"
        elif not category:
            category = _infer_category(desc, debit, credit, is_kushki_credit=False)
        running_balance += (credit - debit)
        category_acc[category]["debits"] += debit
        category_acc[category]["credits"] += credit
        category_acc[category]["count"] += 1
        ws_moves.append([
            d.day if d else None,
            d if d else _safe_str(mv.get("date"), ""),
            typ,
            desc,
            category,
            debit if debit > 0 else None,
            credit if credit > 0 else None,
            round(running_balance, 6),
        ])

    # ---- RESUMEN ----
    ws_summary["A1"] = f"RESUMEN {period_text} – BANREGIO TRES COMAS"
    summary_headers = ["CATEGORÍA", "TOTAL CARGOS", "TOTAL ABONOS", "# MOVS"]
    ws_summary.append(summary_headers)
    _styled_header_row(ws_summary, 2, 1, len(summary_headers))
    for category, data in sorted(category_acc.items(), key=lambda x: x[0]):
        ws_summary.append([
            category,
            round(data["debits"], 6) if data["debits"] else None,
            round(data["credits"], 6) if data["credits"] else None,
            data["count"],
        ])

    # ---- CRUCE KUSHKI ----
    ws_cross["A1"] = f"CRUCE KUSHKI vs BANREGIO – LIQUIDACIONES {period_text}   |   TRES COMAS S.A.P.I. DE C.V."
    total_expected = len([d for d in daily_summary if _num(d.get("net_deposit")) > 0])
    total_matched = len(matched)
    total_diff_rows = len(kvb_differences) + len(unmatched_banregio)
    ws_cross["A2"] = (
        f"{total_expected} depósitos esperados · {total_matched} conciliados · "
        f"{total_diff_rows} diferencias · Diferencia total: ${total_difference:,.2f}"
    )
    cross_headers = [
        "FECHA LIQ.",
        "# TXNS",
        "MONTO BRUTO",
        "COM. KUSHKI + IVA",
        "RR RETENIDO",
        "RR LIBERADO",
        "DEP. NETO KUSHKI",
        "ABONO BANREGIO",
        "DIFERENCIA",
        "ESTADO",
        None,
        "RESUMEN DEL CRUCE",
        None,
    ]
    ws_cross.append(cross_headers)
    _styled_header_row(ws_cross, 3, 1, 10)
    _styled_header_row(ws_cross, 3, 12, 13)

    cross_match_by_date: Dict[str, Dict[str, Any]] = {}
    for row in (matched + kvb_differences):
        date_key = _date_label(row.get("date"))
        if not date_key:
            continue
        candidate = {
            "banregio_amount": round(_num(row.get("banregio_amount")), 6),
            "banregio_date": _date_label(row.get("banregio_date")),
            "difference": round(_num(row.get("difference")), 6),
        }
        current = cross_match_by_date.get(date_key)
        if current is None or candidate["difference"] < current["difference"]:
            cross_match_by_date[date_key] = candidate

    cross_rows = []
    for daily in sorted(daily_summary, key=lambda r: _date_label(r.get("date"))):
        kushki_amount = round(_num(daily.get("net_deposit")), 6)
        if kushki_amount <= 0:
            continue
        date_key = _date_label(daily.get("date"))
        matched_row = cross_match_by_date.get(date_key)
        banregio_amount = round(_num((matched_row or {}).get("banregio_amount")), 6)
        diff = round(abs(kushki_amount - banregio_amount), 6) if banregio_amount > 0 else kushki_amount
        status = "OK" if banregio_amount > 0 and diff <= 0.01 else "Revisar"
        rr_retained = round(_num(daily.get("rr_retained")), 6)
        rr_released = round(_num(daily.get("rr_released")), 6)
        rolling = round(_num(daily.get("rolling_reserve")), 6)
        if rr_retained == 0 and rr_released == 0 and rolling != 0:
            if rolling > 0:
                rr_retained = rolling
            else:
                rr_released = abs(rolling)
        ws_cross.append([
            _parse_date(date_key) or date_key,
            int(_num(daily.get("tx_count"))),
            round(_num(daily.get("gross_amount")), 6),
            round(_num(daily.get("commission")), 6),
            rr_retained,
            rr_released,
            kushki_amount,
            banregio_amount if banregio_amount > 0 else None,
            diff if diff else None,
            status,
            None,
            None,
            None,
        ])
        cross_rows.append({
            "kushki_amount": kushki_amount,
            "banregio_amount": banregio_amount,
            "difference": diff,
            "status": status,
        })

    cross_total_kushki = round(sum(r["kushki_amount"] for r in cross_rows), 6)
    cross_total_banregio = round(sum(r["banregio_amount"] for r in cross_rows), 6)
    cross_days_with_diff = sum(1 for r in cross_rows if r["difference"] > 0.01)
    side_rows = [
        ("Depósitos Kushki", cross_total_kushki),
        ("Abonos Banregio", cross_total_banregio),
        ("Diferencia total", round(abs(cross_total_kushki - cross_total_banregio), 6)),
        ("Días conciliados", total_matched),
        ("Días con diferencia", cross_days_with_diff),
    ]
    side_start = 4
    for i, (label, value) in enumerate(side_rows):
        ws_cross.cell(row=side_start + i, column=12, value=label)
        ws_cross.cell(row=side_start + i, column=13, value=value)

    for ws in [ws_moves, ws_summary, ws_cross]:
        _autowidth(ws)

    filename = f"BANREGIO_{_month_name_upper(process.period_month)}_{process.period_year}_CONCILIADO_v2.xlsx"
    return filename, _save_workbook(wb)
