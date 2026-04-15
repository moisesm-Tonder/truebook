import io
import unittest
from types import SimpleNamespace

import pandas as pd
from openpyxl import load_workbook

from app.services.conciliation_engine import conciliate_kushki_daily
from app.services.excel_exports import build_kushki_export
from app.services.kushki_parser import (
    enrich_kushki_with_fees,
    merge_kushki_results,
    parse_kushki,
)


class KushkiPdfImplementationTests(unittest.TestCase):
    def _build_settlement_excel_bytes(self, rows):
        df = pd.DataFrame(rows)
        buff = io.BytesIO()
        with pd.ExcelWriter(buff, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Detalle de Liquidacion", index=False)
        return buff.getvalue()

    def test_parse_kushki_raw_settlement_computes_pdf_fields(self):
        rows = [
            {
                "Fecha Pago": "2026-02-03",
                "merchant_name": "AFUN",
                "transaction_status": "APPROVED",
                "transaction_type": "SALE",
                "approved_transaction_amount": 1000.0,
                "kushki_commission": 20.0,
                "iva_kushki_commission": 3.2,
                "fraud_retention": 10.0,
                "Liberacion de Fondos": 0.0,
                "Monto Abonar": 966.8,
            },
            {
                "Fecha Pago": "2026-02-03",
                "merchant_name": "AFUN",
                "transaction_status": "APPROVED",
                "transaction_type": "REFUND",
                "approved_transaction_amount": 0.0,
                "kushki_commission": 0.0,
                "iva_kushki_commission": 0.0,
                "fraud_retention": 0.0,
                "Liberacion de Fondos": 0.0,
                "Monto Abonar": -50.0,
            },
            {
                "Fecha Pago": "2026-02-03",
                "merchant_name": "AFUN",
                "transaction_status": "APPROVED",
                "transaction_type": "CHARGEBACK",
                "approved_transaction_amount": 0.0,
                "kushki_commission": 0.0,
                "iva_kushki_commission": 0.0,
                "fraud_retention": 0.0,
                "Liberacion de Fondos": 0.0,
                "Monto Abonar": -30.0,
            },
            {
                "Fecha Pago": "2026-02-03",
                "merchant_name": "AFUN",
                "transaction_status": "APPROVED",
                "transaction_type": "MANUAL",
                "approved_transaction_amount": 0.0,
                "kushki_commission": 0.0,
                "iva_kushki_commission": 0.0,
                "fraud_retention": 0.0,
                "Liberacion de Fondos": 0.0,
                "Monto Abonar": 5.0,
            },
        ]
        content = self._build_settlement_excel_bytes(rows)

        parsed = parse_kushki(content, "TONDER_2026-02-03.xlsx")

        self.assertEqual(len(parsed["daily_summary"]), 1)
        daily = parsed["daily_summary"][0]

        self.assertEqual(daily["date"], "2026-02-03")
        self.assertEqual(daily["tx_count"], 1)
        self.assertAlmostEqual(daily["gross_amount"], 1000.0, places=6)
        self.assertAlmostEqual(daily["refund"], -50.0, places=6)
        self.assertAlmostEqual(daily["chargeback"], -30.0, places=6)
        self.assertAlmostEqual(daily["manual"], 5.0, places=6)
        self.assertAlmostEqual(daily["gross_adjustments"], 0.0, places=6)
        self.assertAlmostEqual(daily["net_deposit"], 891.8, places=6)
        self.assertAlmostEqual(daily["net_verification"], 891.8, places=6)
        self.assertAlmostEqual(daily["validation_diff"], 0.0, places=6)

    def test_merge_and_enrich_with_fees_periods(self):
        kushki_data = {
            "daily_summary": [
                {
                    "date": "2026-02-03",
                    "tx_count": 1,
                    "gross_amount": 1000,
                    "gross_adjustments": 0,
                    "commission": 23.2,
                    "kushki_commission": 20,
                    "iva_kushki_commission": 3.2,
                    "rr_retained": 10,
                    "rr_released": 0,
                    "refund": 0,
                    "chargeback": 0,
                    "void": 0,
                    "manual": 0,
                    "adjustments": 0,
                    "rolling_reserve": 10,
                    "net_deposit": 966.8,
                    "net_verification": 966.8,
                    "validation_diff": 0,
                },
                {
                    "date": "2026-02-05",
                    "tx_count": 1,
                    "gross_amount": 500,
                    "gross_adjustments": 0,
                    "commission": 10,
                    "kushki_commission": 8.62,
                    "iva_kushki_commission": 1.38,
                    "rr_retained": 5,
                    "rr_released": 0,
                    "refund": 0,
                    "chargeback": 0,
                    "void": 0,
                    "manual": 0,
                    "adjustments": 0,
                    "rolling_reserve": 5,
                    "net_deposit": 485,
                    "net_verification": 485,
                    "validation_diff": 0,
                },
            ],
            "merchant_detail": [
                {
                    "date": "2026-02-03",
                    "merchant_name": "AFUN",
                    "tx_count": 1,
                    "gross_amount": 1000,
                    "gross_adjustments": 0,
                    "commission": 23.2,
                    "kushki_commission": 20,
                    "iva_kushki_commission": 3.2,
                    "rr_retained": 10,
                    "rr_released": 0,
                    "refund": 0,
                    "chargeback": 0,
                    "void": 0,
                    "manual": 0,
                    "adjustments": 0,
                    "rolling_reserve": 10,
                    "net_deposit": 966.8,
                    "net_verification": 966.8,
                    "validation_diff": 0,
                },
                {
                    "date": "2026-02-05",
                    "merchant_name": "NUEVO SIN MAP",
                    "tx_count": 1,
                    "gross_amount": 500,
                    "gross_adjustments": 0,
                    "commission": 10,
                    "kushki_commission": 8.62,
                    "iva_kushki_commission": 1.38,
                    "rr_retained": 5,
                    "rr_released": 0,
                    "refund": 0,
                    "chargeback": 0,
                    "void": 0,
                    "manual": 0,
                    "adjustments": 0,
                    "rolling_reserve": 5,
                    "net_deposit": 485,
                    "net_verification": 485,
                    "validation_diff": 0,
                },
            ],
            "total_net_deposit": 1451.8,
        }

        # also validate merge path with one item
        merged = merge_kushki_results([kushki_data])

        fees_data = {
            "daily_breakdown": [
                {"date": "2026-02-01", "merchant_name": "Afun Mexico", "acquirer": "Kushki", "fee_amount": 10},
                {"date": "2026-02-02", "merchant_name": "Afun Mexico", "acquirer": "Kushki", "fee_amount": 20},
                {"date": "2026-02-03", "merchant_name": "Afun Mexico", "acquirer": "Bitso", "fee_amount": 5},
                {"date": "2026-02-04", "merchant_name": "Afun Mexico", "acquirer": "Kushki", "fee_amount": 7},
                {"date": "2026-02-05", "merchant_name": "Afun Mexico", "acquirer": "Kushki", "fee_amount": 8},
            ]
        }

        enriched = enrich_kushki_with_fees(merged, fees_data, 2026, 2)

        by_date = {r["date"]: r for r in enriched["daily_summary"]}
        self.assertAlmostEqual(by_date["2026-02-03"]["tonder_commission"], 35.0, places=6)  # all concepts 1-3
        self.assertAlmostEqual(by_date["2026-02-05"]["tonder_commission"], 15.0, places=6)  # all concepts 4-5

        afun_row = next(r for r in enriched["merchant_detail"] if r["merchant_name"] == "AFUN")
        no_map_row = next(r for r in enriched["merchant_detail"] if r["merchant_name"] == "NUEVO SIN MAP")
        self.assertAlmostEqual(afun_row["tonder_commission"], 30.0, places=6)  # only kushki concept 1-3
        self.assertAlmostEqual(no_map_row["tonder_commission"], 0.0, places=6)
        self.assertIn("NUEVO SIN MAP", enriched.get("unmapped_merchants", []))

    def test_conciliate_kushki_daily_uses_full_formula(self):
        kushki_result = {
            "daily_summary": [
                {
                    "date": "2026-02-03",
                    "tx_count": 1,
                    "gross_amount": 200,
                    "gross_adjustments": -20,
                    "commission": 10,
                    "rr_retained": 5,
                    "rr_released": 0,
                    "refund": -30,
                    "chargeback": -10,
                    "void": -5,
                    "manual": 0,
                    "rolling_reserve": 5,
                    "net_deposit": 120,
                },
                {
                    "date": "2026-02-04",
                    "tx_count": 1,
                    "gross_amount": 200,
                    "gross_adjustments": -20,
                    "commission": 10,
                    "rr_retained": 5,
                    "rr_released": 0,
                    "refund": -30,
                    "chargeback": -10,
                    "void": -5,
                    "manual": 0,
                    "rolling_reserve": 5,
                    "net_deposit": 119,
                },
            ]
        }

        result = conciliate_kushki_daily(kushki_result)

        self.assertEqual(len(result["matched"]), 1)
        self.assertEqual(len(result["differences"]), 1)
        self.assertAlmostEqual(result["matched"][0]["computed_net"], 120.0, places=6)
        self.assertAlmostEqual(result["differences"][0]["difference"], 1.0, places=6)

    def test_build_kushki_export_creates_pdf_structure(self):
        daily_summary = [
            {
                "date": "2026-02-03",
                "tx_count": 1,
                "gross_amount": 1000,
                "gross_adjustments": 0,
                "commission": 23.2,
                "kushki_commission": 20,
                "iva_kushki_commission": 3.2,
                "rr_retained": 10,
                "rr_released": 0,
                "refund": 0,
                "chargeback": 0,
                "void": 0,
                "manual": 0,
                "adjustments": 0,
                "rolling_reserve": 10,
                "net_deposit": 966.8,
                "tonder_commission": 30,
                "tonder_iva": 4.8,
                "tonder_total": 34.8,
            }
        ]
        merchant_detail = [
            {
                "date": "2026-02-03",
                "merchant_name": "AFUN",
                "tx_count": 1,
                "gross_amount": 1000,
                "gross_adjustments": 0,
                "commission": 23.2,
                "kushki_commission": 20,
                "iva_kushki_commission": 3.2,
                "rr_retained": 10,
                "rr_released": 0,
                "refund": 0,
                "chargeback": 0,
                "void": 0,
                "manual": 0,
                "adjustments": 0,
                "rolling_reserve": 10,
                "net_deposit": 966.8,
                "tonder_commission": 30,
                "tonder_iva": 4.8,
                "tonder_total": 34.8,
            }
        ]

        process = SimpleNamespace(period_month=2, period_year=2026)
        kushki_result = SimpleNamespace(daily_summary=daily_summary, merchant_detail=merchant_detail)

        filename, payload = build_kushki_export(process, kushki_result)
        self.assertEqual(filename, "KUSHKI_FEBRERO_2026_v4.xlsx")
        self.assertGreater(len(payload), 0)

        wb = load_workbook(io.BytesIO(payload), data_only=False)
        self.assertEqual(wb.sheetnames, ["Detalle por Merchant", "Resumen Diario", "Pivot por Merchant"])

        ws_detail = wb["Detalle por Merchant"]
        ws_daily = wb["Resumen Diario"]
        ws_pivot = wb["Pivot por Merchant"]

        self.assertEqual(ws_detail["A3"].value, "Fecha Liq.")
        self.assertEqual(ws_daily["A2"].value, "Fecha Liq.")
        self.assertEqual(ws_pivot["A2"].value, "Merchant")


if __name__ == "__main__":
    unittest.main()
